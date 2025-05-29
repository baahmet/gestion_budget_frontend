from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QLabel, QTableWidget, QTableWidgetItem,
                             QHeaderView, QMessageBox, QHBoxLayout, QPushButton, QComboBox,
                             QLineEdit, QGraphicsDropShadowEffect)
from PyQt5.QtCore import Qt, QPropertyAnimation, QEasingCurve
from PyQt5.QtGui import QColor, QFont, QBrush
from services.audit_service import get_audit_logs
from services.auth_service import AuthService
import datetime
import math


class JournalAuditWidget(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Journal d'Audit")
        self.audit_logs = []
        self.filtered_logs = []
        self.current_page = 1
        self.items_per_page = 10
        self.setup_ui()
        self.load_audit_logs()
        self.setup_animations()

    def setup_ui(self):
        # Style global
        self.setStyleSheet("""
            QWidget {
                background-color: #f5f7fa;
                font-family: 'Segoe UI';
            }
            QToolTip {
                background-color: #34495e;
                color: white;
                border: none;
                padding: 5px;
                border-radius: 2px;
            }
        """)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # En-tête
        header = QHBoxLayout()

        title = QLabel("JOURNAL D'AUDIT DES ACTIONS UTILISATEURS")
        title.setStyleSheet("""
            QLabel {
                font-size: 22px;
                font-weight: bold;
                color: #2c3e50;
            }
        """)
        header.addWidget(title)
        header.addStretch()
        layout.addLayout(header)

        # Barre de filtres
        filter_bar = QHBoxLayout()

        # Filtre par utilisateur
        self.user_filter = QComboBox()
        self.user_filter.addItem("Tous les utilisateurs", "tous")
        self.user_filter.setStyleSheet("""
            QComboBox {
                padding: 5px;
                border: 1px solid #bdc3c7;
                border-radius: 3px;
                background-color: white;
                min-width: 150px;
            }
            QComboBox::drop-down {
                border: none;
                width: 20px;
            }
        """)
        self.user_filter.currentIndexChanged.connect(self.apply_filters)
        filter_bar.addWidget(QLabel("Utilisateur:"))
        filter_bar.addWidget(self.user_filter)

        # Filtre par action
        self.action_filter = QComboBox()
        self.action_filter.addItem("Toutes les actions", "toutes")
        self.action_filter.setStyleSheet("""
            QComboBox {
                padding: 5px;
                border: 1px solid #bdc3c7;
                border-radius: 3px;
                background-color: white;
                min-width: 150px;
            }
            QComboBox::drop-down {
                border: none;
                width: 20px;
            }
        """)
        self.action_filter.currentIndexChanged.connect(self.apply_filters)
        filter_bar.addWidget(QLabel("Action:"))
        filter_bar.addWidget(self.action_filter)

        # Filtre de recherche
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Rechercher...")
        self.search_input.setStyleSheet("""
            QLineEdit {
                padding: 5px;
                border: 1px solid #bdc3c7;
                border-radius: 3px;
                background-color: white;
                min-width: 200px;
            }
        """)
        self.search_input.textChanged.connect(self.apply_filters)
        filter_bar.addWidget(QLabel("Recherche:"))
        filter_bar.addWidget(self.search_input)

        filter_bar.addStretch()

        # Sélecteur d'éléments par page
        self.items_per_page_combo = QComboBox()
        self.items_per_page_combo.addItems(["5", "10", "20", "50", "100"])
        self.items_per_page_combo.setCurrentText("10")
        self.items_per_page_combo.setStyleSheet("""
            QComboBox {
                padding: 5px;
                border: 1px solid #bdc3c7;
                border-radius: 3px;
                background-color: white;
                min-width: 80px;
            }
        """)
        self.items_per_page_combo.currentTextChanged.connect(self.change_items_per_page)
        filter_bar.addWidget(QLabel("Items par page:"))
        filter_bar.addWidget(self.items_per_page_combo)

        # Affichage du nombre de résultats
        self.results_label = QLabel("0 entrées")
        self.results_label.setStyleSheet("color: #7f8c8d;")
        filter_bar.addWidget(self.results_label)

        layout.addLayout(filter_bar)

        # Tableau des logs d'audit
        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["Utilisateur", "Email", "Action", "Date"])
        self.table.setStyleSheet("""
            QTableWidget {
                background-color: white;
                border-radius: 8px;
                gridline-color: #ecf0f1;
                font-size: 13px;
                selection-background-color: #3498db;
                selection-color: white;
            }
            QHeaderView::section {
                background-color: #34495e;
                color: white;
                padding: 10px;
                border: none;
                font-weight: bold;
            }
            QTableWidget::item {
                padding: 10px;
            }
            QTableWidget::item:selected {
                background-color: #e0f2fe;
                color: #2c3e50;
            }
        """)

        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.setShowGrid(False)

        # Effet d'ombre
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(15)
        shadow.setColor(QColor(0, 0, 0, 30))
        shadow.setOffset(0, 3)
        self.table.setGraphicsEffect(shadow)

        layout.addWidget(self.table)

        # Pagination
        self.pagination_widget = QWidget()
        pagination_layout = QHBoxLayout(self.pagination_widget)
        pagination_layout.setContentsMargins(0, 10, 0, 10)
        pagination_layout.setSpacing(5)

        self.prev_btn = QPushButton("◀")
        self.prev_btn.setStyleSheet("""
            QPushButton {
                background-color: #bdc3c7;
                color: #2c3e50;
                border: none;
                padding: 5px 10px;
                border-radius: 3px;
                min-width: 30px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #95a5a6;
            }
            QPushButton:disabled {
                background-color: #ecf0f1;
                color: #95a5a6;
            }
        """)
        self.prev_btn.setCursor(Qt.PointingHandCursor)
        self.prev_btn.clicked.connect(self.go_to_previous_page)
        self.prev_btn.setToolTip("Page précédente")
        pagination_layout.addWidget(self.prev_btn)

        self.page_buttons_layout = QHBoxLayout()
        self.page_buttons_layout.setSpacing(5)
        pagination_layout.addLayout(self.page_buttons_layout)

        self.next_btn = QPushButton("▶")
        self.next_btn.setStyleSheet("""
            QPushButton {
                background-color: #bdc3c7;
                color: #2c3e50;
                border: none;
                padding: 5px 10px;
                border-radius: 3px;
                min-width: 30px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #95a5a6;
            }
            QPushButton:disabled {
                background-color: #ecf0f1;
                color: #95a5a6;
            }
        """)
        self.next_btn.setCursor(Qt.PointingHandCursor)
        self.next_btn.clicked.connect(self.go_to_next_page)
        self.next_btn.setToolTip("Page suivante")
        pagination_layout.addWidget(self.next_btn)

        pagination_layout.addStretch()

        self.page_info_label = QLabel()
        self.page_info_label.setStyleSheet("color: #7f8c8d;")
        pagination_layout.addWidget(self.page_info_label)

        layout.addWidget(self.pagination_widget)

        # Barre de statut
        status_bar = QHBoxLayout()

        # Date de dernière mise à jour
        self.update_label = QLabel(f"Dernière mise à jour: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}")
        self.update_label.setStyleSheet("color: #7f8c8d; font-style: italic;")
        status_bar.addWidget(self.update_label)

        status_bar.addStretch()

        # Information sur le rôle utilisateur
        user_role = AuthService.get_user_role()

        role_label = QLabel(f"Rôle actuel: {user_role.capitalize()}")
        role_label.setStyleSheet("""
            color: #34495e;
            font-weight: bold;
            padding: 5px 10px;
            background-color: #ecf0f1;
            border-radius: 3px;
        """)
        status_bar.addWidget(role_label)

        # Bouton de rafraîchissement
        refresh_btn = QPushButton("🔄 Rafraîchir")
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #ecf0f1;
                color: #2c3e50;
                border: none;
                padding: 5px 10px;
                border-radius: 3px;
            }
            QPushButton:hover {
                background-color: #d5dbdb;
            }
        """)
        refresh_btn.setCursor(Qt.PointingHandCursor)
        refresh_btn.clicked.connect(self.refresh_data)
        status_bar.addWidget(refresh_btn)

        # Bouton d'exportation
        export_btn = QPushButton("📊 Exporter")
        export_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                padding: 5px 10px;
                border-radius: 3px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        export_btn.setCursor(Qt.PointingHandCursor)
        export_btn.clicked.connect(self.export_data)
        export_btn.setToolTip("Exporter les logs d'audit au format CSV")
        status_bar.addWidget(export_btn)

        layout.addLayout(status_bar)

    def load_audit_logs(self):
        result = get_audit_logs()
        if result["success"]:
            self.audit_logs = result["data"]
            self.filtered_logs = self.audit_logs

            # Extraire les utilisateurs et actions uniques pour les filtres
            users = set()
            actions = set()
            for log in self.audit_logs:
                users.add(log.get("utilisateur_nom"))
                actions.add(log.get("action", "N/A"))

            # Mettre à jour les combobox de filtres
            current_user = self.user_filter.currentText()
            current_action = self.action_filter.currentText()

            self.user_filter.clear()
            self.user_filter.addItem("Tous les utilisateurs", "tous")
            for user in sorted(users):
                self.user_filter.addItem(user, user)

            self.action_filter.clear()
            self.action_filter.addItem("Toutes les actions", "toutes")
            for action in sorted(actions):
                self.action_filter.addItem(action, action)

            # Restaurer les sélections précédentes si possible
            user_index = self.user_filter.findText(current_user)
            if user_index > 0:
                self.user_filter.setCurrentIndex(user_index)

            action_index = self.action_filter.findText(current_action)
            if action_index > 0:
                self.action_filter.setCurrentIndex(action_index)

            self.apply_filters()
        else:
            self.show_error_message("Erreur", result["message"])

    def apply_filters(self):
        # Récupérer les valeurs des filtres
        user_filter = self.user_filter.currentData()
        action_filter = self.action_filter.currentData()
        search_text = self.search_input.text().lower()

        # Filtrer les logs
        self.filtered_logs = []
        for log in self.audit_logs:
            # Filtre par utilisateur
            if user_filter != "tous" and log.get("utilisateur_nom", "Inconnu") != user_filter:
                continue

            # Filtre par action
            if action_filter != "toutes" and log.get("action", "N/A") != action_filter:
                continue

            # Filtre par texte de recherche
            searchable_text = (
                    log.get("utilisateur_nom", "").lower() +
                    log.get("utilisateur_email", "").lower() +
                    log.get("action", "").lower() +
                    log.get("date_heure", "").lower()
            )

            if search_text and search_text not in searchable_text:
                continue

            self.filtered_logs.append(log)

        self.current_page = 1  # Reset to first page when filters change
        self.update_table()
        self.update_pagination()
        self.results_label.setText(f"{len(self.filtered_logs)} entrée(s) trouvée(s)")

    def change_items_per_page(self, text):
        self.items_per_page = int(text)
        self.current_page = 1  # Reset to first page when items per page changes
        self.update_table()
        self.update_pagination()

    def update_table(self):
        start_index = (self.current_page - 1) * self.items_per_page
        end_index = start_index + self.items_per_page
        paginated_logs = self.filtered_logs[start_index:end_index]

        self.table.setRowCount(len(paginated_logs))

        for i, log in enumerate(paginated_logs):
            nom = log.get("utilisateur_nom", "Inconnu")
            email = log.get("utilisateur_email", "Inconnu")
            action = log.get("action", "N/A")
            date_heure = log.get("date_heure", "")[:19]  # Format ISO

            # Utilisateur
            nom_item = QTableWidgetItem(nom)
            self.table.setItem(i, 0, nom_item)

            # Email
            email_item = QTableWidgetItem(email)
            self.table.setItem(i, 1, email_item)

            # Action (avec couleur selon le type)
            action_item = QTableWidgetItem(action)

            # Coloration selon le type d'action
            if "création" in action.lower() or "ajout" in action.lower():
                action_item.setForeground(QColor("#27ae60"))  # Vert
            elif "suppression" in action.lower() or "supprimer" in action.lower():
                action_item.setForeground(QColor("#e74c3c"))  # Rouge
            elif "modification" in action.lower() or "mise à jour" in action.lower():
                action_item.setForeground(QColor("#f39c12"))  # Orange
            elif "connexion" in action.lower() or "login" in action.lower():
                action_item.setForeground(QColor("#3498db"))  # Bleu

            self.table.setItem(i, 2, action_item)

            # Date (formatée)
            date_item = QTableWidgetItem(date_heure)
            self.table.setItem(i, 3, date_item)

            # Alternance des couleurs de ligne
            if i % 2 == 0:
                for j in range(self.table.columnCount()):
                    if self.table.item(i, j):
                        self.table.item(i, j).setBackground(QColor("#f8f9fa"))

    def update_pagination(self):
        # Clear existing page buttons
        for i in reversed(range(self.page_buttons_layout.count())):
            widget = self.page_buttons_layout.itemAt(i).widget()
            if widget is not None:
                widget.deleteLater()

        total_pages = math.ceil(len(self.filtered_logs) / self.items_per_page) or 1

        # Always show first page button
        self.add_page_button(1)

        # Show ellipsis if needed
        if self.current_page > 3:
            ellipsis = QLabel("...")
            ellipsis.setStyleSheet("color: #7f8c8d;")
            self.page_buttons_layout.addWidget(ellipsis)

        # Show current page and neighbors
        start_page = max(2, self.current_page - 1)
        end_page = min(total_pages - 1, self.current_page + 1)

        for page in range(start_page, end_page + 1):
            self.add_page_button(page)

        # Show ellipsis if needed
        if self.current_page < total_pages - 2:
            ellipsis = QLabel("...")
            ellipsis.setStyleSheet("color: #7f8c8d;")
            self.page_buttons_layout.addWidget(ellipsis)

        # Always show last page button if there's more than one page
        if total_pages > 1:
            self.add_page_button(total_pages)

        # Update page info label
        start_item = (self.current_page - 1) * self.items_per_page + 1
        end_item = min(self.current_page * self.items_per_page, len(self.filtered_logs))
        total_items = len(self.filtered_logs)
        self.page_info_label.setText(f"Affichage de {start_item}-{end_item} sur {total_items}")

        # Enable/disable navigation buttons
        self.prev_btn.setEnabled(self.current_page > 1)
        self.next_btn.setEnabled(self.current_page < total_pages)

    def add_page_button(self, page):
        btn = QPushButton(str(page))
        btn.setCheckable(True)
        btn.setChecked(page == self.current_page)
        btn.setStyleSheet("""
            QPushButton {
                background-color: %s;
                color: %s;
                border: none;
                padding: 5px 10px;
                border-radius: 3px;
                min-width: 30px;
            }
            QPushButton:hover {
                background-color: #95a5a6;
                color: white;
            }
            QPushButton:checked {
                background-color: #3498db;
                color: white;
            }
        """ % ("#ecf0f1" if page != self.current_page else "#3498db",
              "#2c3e50" if page != self.current_page else "white"))
        btn.setCursor(Qt.PointingHandCursor)
        btn.clicked.connect(lambda: self.go_to_page(page))
        self.page_buttons_layout.addWidget(btn)

    def go_to_page(self, page):
        self.current_page = page
        self.update_table()
        self.update_pagination()

    def go_to_previous_page(self):
        if self.current_page > 1:
            self.go_to_page(self.current_page - 1)

    def go_to_next_page(self):
        total_pages = math.ceil(len(self.filtered_logs) / self.items_per_page)
        if self.current_page < total_pages:
            self.go_to_page(self.current_page + 1)

    def setup_animations(self):
        # Animation d'apparition du tableau
        self.anim = QPropertyAnimation(self.table, b"windowOpacity")
        self.anim.setDuration(500)
        self.anim.setStartValue(0)
        self.anim.setEndValue(1)
        self.anim.setEasingCurve(QEasingCurve.OutCubic)
        self.anim.start()

    def refresh_data(self):
        self.load_audit_logs()
        self.update_label.setText(f"Dernière mise à jour: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}")

        # Animation de rafraîchissement
        anim = QPropertyAnimation(self.table, b"windowOpacity")
        anim.setDuration(300)
        anim.setStartValue(0.5)
        anim.setEndValue(1)
        anim.setEasingCurve(QEasingCurve.OutCubic)
        anim.start()

    def export_data(self):
        try:
            from PyQt5.QtWidgets import QFileDialog
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            # Demander à l'utilisateur où enregistrer le fichier
            filename, _ = QFileDialog.getSaveFileName(
                self, "Exporter les logs d'audit",
                f"audit_logs_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Fichiers Excel (*.xlsx)"
            )

            if not filename:
                return

            # Créer un nouveau classeur Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Journal d'audit"

            # Styles
            header_font = Font(name='Calibri', bold=True, color="FFFFFF", size=12)  # Police blanche
            header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")  # Fond sombre
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            data_font = Font(name='Calibri', size=11)
            data_alignment = Alignment(horizontal="left", vertical="center")

            # Couleurs pour les types d'action
            action_colors = {
                "création": "27AE60",  # vert
                "ajout": "27AE60",
                "suppression": "E74C3C",  # rouge
                "supprimer": "E74C3C",
                "modification": "F39C12",  # orange
                "mise à jour": "F39C12",
                "connexion": "3498DB",  # bleu
                "login": "3498DB"
            }

            # En-têtes
            headers = ["Utilisateur", "Email", "Action", "Date"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # Données
            for row_num, log in enumerate(self.filtered_logs, 2):
                # Récupérer les données
                nom = log.get("utilisateur_nom", "Inconnu")
                email = log.get("utilisateur_email", "Inconnu")
                action = log.get("action", "N/A")
                date_heure = log.get("date_heure", "")[:19]  # Format ISO

                # Écrire les données
                ws.cell(row=row_num, column=1, value=nom).font = data_font
                ws.cell(row=row_num, column=2, value=email).font = data_font
                action_cell = ws.cell(row=row_num, column=3, value=action)
                ws.cell(row=row_num, column=4, value=date_heure).font = data_font

                # Appliquer le style à toute la ligne
                for col_num in range(1, 5):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.alignment = data_alignment
                    cell.border = thin_border

                    # Alternance des couleurs de fond
                    if row_num % 2 == 0:
                        cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")

                # Colorer la cellule d'action selon le type
                action_lower = action.lower()
                for key, color in action_colors.items():
                    if key in action_lower:
                        action_cell.font = Font(name='Calibri', bold=True, color=color, size=11)
                        break

            # Ajuster la largeur des colonnes
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter  # Get the column name
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = adjusted_width

            # Ajouter un filtre automatique
            ws.auto_filter.ref = f"A1:D{len(self.filtered_logs) + 1}"

            # Ajouter un titre et des métadonnées
            ws['A1'].value = "JOURNAL D'AUDIT DES ACTIONS UTILISATEURS"
            ws.merge_cells('A1:D1')
            title_cell = ws['A1']
            title_cell.font = Font(name='Calibri', bold=True, size=14, color="2C3E50")
            title_cell.alignment = Alignment(horizontal="center")

            # Enregistrer le fichier
            wb.save(filename)

            self.show_success_message("Export réussi",
                                      f"Les logs d'audit ont été exportés avec succès dans le fichier:\n{filename}")
        except Exception as e:
            self.show_error_message("Erreur d'exportation", f"Une erreur est survenue lors de l'exportation: {str(e)}")



    def show_error_message(self, title, message):
        msg = QMessageBox(self)
        msg.setIcon(QMessageBox.Critical)
        msg.setWindowTitle(title)
        msg.setText(message)
        msg.setStyleSheet("""
            QMessageBox {
                background-color: white;
            }
            QMessageBox QLabel {
                color: #e74c3c;
                font-size: 14px;
            }
            QMessageBox QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                padding: 5px 15px;
                border-radius: 3px;
                min-width: 60px;
            }
            QMessageBox QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        msg.exec_()

    def show_success_message(self, title, message):
        msg = QMessageBox(self)
        msg.setIcon(QMessageBox.Information)
        msg.setWindowTitle(title)
        msg.setText(message)
        msg.setStyleSheet("""
            QMessageBox {
                background-color: white;
            }
            QMessageBox QLabel {
                color: #27ae60;
                font-size: 14px;
            }
            QMessageBox QPushButton {
                background-color: #27ae60;
                color: white;
                border: none;
                padding: 5px 15px;
                border-radius: 3px;
                min-width: 60px;
            }
            QMessageBox QPushButton:hover {
                background-color: #2ecc71;
            }
        """)
        msg.exec_()